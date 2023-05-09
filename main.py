from simple_bank_korea.kb import get_transactions
import openpyxl
import requests
import time
from var import *

#Data from Excel files
wsList = []
trList = []
trCount = 0

def get_banktransactions():
    transaction_list = get_transactions(
        bank_num = BANK_NUM,        # 계좌번호
        birthday = BIRTHDAY,        # 사업자번호
        password = PASSWORD,        # 계좌비밀번호
        days = 1,                   # 조회할 기간(기본 30일)
        PHANTOM_PATH = PHANTOMPATH
    )

    for trs in transaction_list:
        if(trs['amount'] > 0):
            trList.append((trs['date'].strftime("%Y/%m/%d %r"), trs['amount'], trs['transaction_by']))
            #print(trs['date'], format(trs['amount'], ','), trs['transaction_by'])
        else:
            continue
    
    global trCount
    #print(trCount)

    if(len(trList) > trCount):
        trCount = len(trList)

        for tr in trList:
            for ws in wsList:
                if(tr[1] == ws[3]):
                    send_LINENotification(tr, ws)
                    del ws
            del tr
    else:
        return

def get_excelinformations():
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    worksheet = workbook[SHEET_NAME]

    #print(worksheet.title)

    for row in worksheet.iter_rows(min_row = 4, max_col = 7, values_only=True):
        if(row[1]==None):
            continue
        else:
            wsList.append((row[1], row[2], row[3], row[6]))


def send_LINENotification(trans, works):
    MESSAGE = "\n"
    #MESSAGE = MESSAGE + trans[0] + "\n" + trans[2] + "에서 "
    MESSAGE = MESSAGE + trans[0] + "\n" + works[1] + "에서\n" + works[0] + "과제 " + works[2] + "분\n" + str(format(trans[1], ',')) + "원이 입금 되었습니다.\n"
    print(MESSAGE)

#    response = requests.post(
#        TARGET_URL,
#        headers={'Authorization': 'Bearer ' + TOKEN},
#        data = {
#            'message': MESSAGE
#        }
#    )

#    print(response.text)


if __name__ == '__main__':
    # Get 미수금 data from excel sheet
    get_excelinformations()

    #while(True):
    get_banktransactions()
    #   time.sleep(3600)
